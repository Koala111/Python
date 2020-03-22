from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side
import os
'''
修改后的文件后缀可能打不开
报错zipfile.BadZipFile: File is not a zip file
只能打开xlsx格式的文件

'''
os.chdir(r'F:\demo') #进入指定的目录
filename='1.xlsx'
workbook = load_workbook(filename)
print(workbook.sheetnames)   #查看1个文档里面有哪几个表格
sheet = workbook['代雄'] #查看一个sheet表
print(sheet.dimensions) #打印出这个表的维度
cell = sheet['A2']
print(cell.value,cell.row,cell.column,cell.coordinate)
'''
sheet = workbook.active #如果就只有一个表的话这样表示
cell = sheet['A2']
cell = sheet.cell(row=1,column=1) #指定行列
cells = sheet['A1:A5'] #元组，以for循环输出
print(cells)
print(cell.value,cell.row,cell.column,cell.coordinate)  #读取某个格子的内容,行，列，坐标
for row in sheet.rows:
	print(row)
for row in  sheet.iter_rows(min_row=2,max_row=3,min_col=1,max_col=2):
	print(row)

font = cell.font
print(font.name,font.size,font.bold,font.italic)

alignment = Alignment(horizontal='center',vertical='center',text_rotation=45)
cell.alignment = alignment

pattern_fill = PatternFill(fill_type='solid',fgColor='99ccff') #渐变色
cell.fill=pattern_fill
sheet.merge_cells('C1:D2') #合并单元格
sheet.unmerge_cells('C1:D2') #取消合并单元格
sheet.merge_cells(start_row = 7,start_column=1,end_row=8,end_column=4)
'''
side = Side(style='thin',color = 'FF000000') 
border = Border(left=side,right=side,top=side,bottom=side)
cell.border = border

#font = Font(name='思源黑体 Regular',size=12,bold=True,italic=True,color='FF0000')
#cell.font = font
#workbook.save(filename='这是一个表格.xlsx')